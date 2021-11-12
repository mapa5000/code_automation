import pandas as pd 
import numpy as np 
import openpyxl 
import os

print ("Starting")
#path = r"P:\Houston Gas Drafting\GIS\Corrections\Submitted\2021\Marco Portillo\pending EOP Data Exploration\test\BEx_RouteSmart_pt_Facet_Join.xlsx"
path = "data.xlsx"

#reading df
dataframe = pd.read_excel(path)

dataframe['House'] = dataframe['AddNum'].astype(str).map(lambda x:x.split('.')[0])
dataframe['Full_name'] = dataframe['Last_Name']+', '+dataframe['First_Name']
dataframe['Full_add'] = dataframe['StName']+', '+dataframe['StType']
df = dataframe[['House','Full_add','Full_name', 'Side','FACET_NAME']]

#sorting values 1)facet 2)last_name
df.sort_values(by=['FACET_NAME',])

#creating files from facet unique values
split_values = df['FACET_NAME'].unique()
#print('Unique Values: {}'.format(split_values))

dfs = []
values=[]
for value in split_values: 
    df1 = df[df['FACET_NAME']==value] 
    dfs.append(df1)
    values.append(value)

sheets = []
sheetmax = 20
book = 0;

for  dfi in dfs:
    i = 0;
    sh = 1
    df_empty_list=[]
    #print('Book: ' + str(values[book]))

    df_empty = pd.DataFrame({})
    for idx in dfi.index:
        i = i + 1
        entry = df.loc[[idx]]
        
        df_empty=df_empty.append([entry] )
        if ((i % sheetmax)==0 or (i == len(dfi))):
            sh= sh + 1   
            df_empty_list.append(df_empty)
            df_empty = pd.DataFrame({})

    writer = pd.ExcelWriter(r'facet_container2/FACET_'+str(values[book])+'.xlsx', engine='xlsxwriter')   
    sh = 1   

    for dfj in df_empty_list:
        dfj.to_excel(writer, sheet_name='Sheet'+str(sh))
        sh = sh + 1

    writer.save()
    book = book + 1 

os.chdir(r"facet_container2/")
os.getcwd()
os.listdir()

# iterating on excel facet list
list_wb = os.listdir()
# iterating on excel facet list
list_wb = os.listdir()

template = r"../template.xlsx"
wb_template = openpyxl.load_workbook(template)

for wb in list_wb:
    if ".xls" in wb:
        wb_obj = openpyxl.load_workbook(wb)
        sheetIdx = 0
        newPath = r"../final_container/" + wb + ".xlsx"
        wb_template.save(newPath)
        wb_new = openpyxl.load_workbook(newPath)
        source = wb_new.active

        # iterating between worksheets in each workbook
        for sheet in wb_obj.worksheets:
            sheetIdx = sheetIdx  + 1

            customer_fullname_row = []
            street_name_row = []
            street_number_row = []
            for irow in range(20):
                #print('irow :' + str(irow))
                customer_fullname_row.append( sheet.cell(row = irow + 2,   column = 5))
                street_name_row.append(  sheet.cell(row = irow + 2,  column = 4))
                street_number_row.append(sheet.cell(row = irow + 2,  column = 3))


            if sheetIdx == 1:
                sheet_template  = source
            else:
                sheet_template = wb_new.copy_worksheet(source)
                sheet_template.title = "Sheet" + str(sheetIdx)


            for irow in range(20):            
                sheet_template['B' + str(irow + 5 )].value  = customer_fullname_row[irow].value
                sheet_template['D' + str(irow + 5 )].value = street_name_row[irow].value
                sheet_template['C' + str(irow + 5 )].value  = street_number_row[irow].value

        wb_new.save(newPath)

print('finished...so far !!!')